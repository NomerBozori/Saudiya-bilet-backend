import hashlib
import hmac
import html
import io
import json
import logging
import os
import re
import secrets
import time
from urllib.parse import parse_qsl, quote
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta, timezone
from collections import defaultdict

import httpx
from aiogram import Bot, Dispatcher
from aiogram.types import BufferedInputFile, Update
from fastapi import Depends, FastAPI, File, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

import database as db
import travelpayouts as tp
from bot_handlers import router as bot_router
from config import settings
from keyboards import admin_order_keyboard
from order_actions import confirm_order as confirm_order_action
from order_actions import reject_order as reject_order_action

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("saudiya-bilet")

# ==================== XAVFSIZLIK: BRUTE-FORCE HIMOYASI ====================
# Admin loginga qiladigan urinishlarni cheklaydi (5 ta urinish, 15 daqiqa blokirovka)
_brute_force_store: dict[str, dict] = defaultdict(lambda: {"attempts": 0, "locked_until": 0.0})
BRUTE_FORCE_MAX_ATTEMPTS = 5
BRUTE_FORCE_LOCKOUT_SECONDS = 15 * 60  # 15 daqiqa


def _check_brute_force(ip: str) -> bool:
    """True qaytaradi agar IP bloklangan bo'lsa."""
    now = time.time()
    entry = _brute_force_store[ip]
    if entry["locked_until"] > now:
        return True
    if entry["attempts"] >= BRUTE_FORCE_MAX_ATTEMPTS:
        entry["locked_until"] = now + BRUTE_FORCE_LOCKOUT_SECONDS
        entry["attempts"] = 0
        log.warning(f"Brute-force hujumi bloklandi: {ip}")
        return True
    return False


def _record_failed_attempt(ip: str) -> None:
    """Muvaffaqatsiz urinishni qayd etadi."""
    entry = _brute_force_store[ip]
    entry["attempts"] += 1
    if entry["attempts"] >= BRUTE_FORCE_MAX_ATTEMPTS:
        entry["locked_until"] = time.time() + BRUTE_FORCE_LOCKOUT_SECONDS
        log.warning(f"Brute-force hujumi bloklandi: {ip} - {BRUTE_FORCE_MAX_ATTEMPTS} urinish")
    else:
        remaining = BRUTE_FORCE_MAX_ATTEMPTS - entry["attempts"]
        log.info(f"Admin login muvaffaqatsiz: {ip} - {remaining} urinish qoldi")


def _clear_brute_force(ip: str) -> None:
    """Muvaffaqatli kirgandan so'ng cheklovlarni olib tashlaydi."""
    if ip in _brute_force_store:
        del _brute_force_store[ip]


def _get_client_ip(request: Request) -> str:
    """Mijoz IP manzilini olish (X-Forwarded-For yoki client.host)."""
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


# ==================== XAVFSIZLIK: FAYL TURINI TEKSHIRISH ====================
ALLOWED_PAYMENT_CONTENT_TYPES = {
    "image/jpeg",
    "image/png",
    "image/gif",
    "image/webp",
    "application/pdf",
}
ALLOWED_PAYMENT_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".pdf"}
MAX_PAYMENT_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


def _validate_payment_file(filename: str, content: bytes, content_type: str) -> None:
    """To'lov cheki faylini tekshiradi: tur, hajim, path traversal."""
    # 1. Hajim tekshirish
    if len(content) > MAX_PAYMENT_FILE_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"Fayl hajmi juda katta. Maksimal {MAX_PAYMENT_FILE_SIZE // (1024*1024)} MB ruxsat etiladi."
        )

    # 2. Fayl nomi path traversal himoyasi
    safe_name = os.path.basename(filename)
    if not safe_name or safe_name != filename or "/" in filename or "\\" in filename or ".." in filename:
        raise HTTPException(status_code=400, detail="Fayl nomi noto'g'ri")

    # 3. Kengaytma tekshirish
    ext = os.path.splitext(safe_name)[1].lower()
    if ext not in ALLOWED_PAYMENT_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Ruxsat etilgan fayl turlari: {', '.join(sorted(ALLOWED_PAYMENT_EXTENSIONS))}"
        )

    # 4. Content-Type tekshirish (xavfsizlik uchun)
    safe_ct = (content_type or "").split(";")[0].strip().lower()
    if safe_ct and safe_ct not in ALLOWED_PAYMENT_CONTENT_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Ruxsat etilgan fayl turlari: {', '.join(sorted(ALLOWED_PAYMENT_CONTENT_TYPES))}"
        )

# Joriy build versiyasi — deploy yangilanganini tekshirish uchun (/api/version)
APP_BUILD = "v14"
APP_BUILD_FEATURES = [
    "ixcham viza kartochkalari va operator tugmalari",
    "ixcham qidiruv oqimi",
    "viza arizalari va admin boshqaruvi",
    "narx tushishi obunasi va Telegram xabari",
    "avto-post 3-35 kun",
    "top-deals avto tavsiyalar",
    "arzon narxlar taqvimi",
    "boarding pass",
    "3D karta + nusxalash",
    "admin: o'chirish/tozalash/excel/CBU",
]

VISA_TYPES = {
    "tourist_multi": "1 yillik Multi Turistik Viza",
    "umrah_nusuk": "Rasmiy Umra Vizasi (Nusuk)",
}
VISA_STATUSES = {"new", "processing", "approved", "rejected"}
PASSPORT_RE = re.compile(r"^[A-Z0-9]{5,20}$")
IATA_RE = re.compile(r"^[A-Z0-9]{3}$")
MAX_ALERT_RANGE_DAYS = 60
TELEGRAM_INIT_DATA_MAX_AGE = 24 * 60 * 60

# Aviakompaniyalarning RASMIY saytlari — admin xabaridagi "Xavfsiz xarid" havolasi uchun.
# Kalit: IATA kodi; qiymat: (rasmiy to'liq nom, rasmiy sayt URL).
AIRLINE_OFFICIAL_SITES: dict[str, tuple[str, str]] = {
    "HY": ("UZBEKISTAN AIRWAYS", "https://www.uzairways.com"),
    "C6": ("CENTRUM AIR", "https://centrum-air.com"),
    "SV": ("SAUDIA", "https://www.saudia.com"),
    "TK": ("TURKISH AIRLINES", "https://www.turkishairlines.com"),
    "FZ": ("FLYDUBAI", "https://www.flydubai.com"),
    "G9": ("AIR ARABIA", "https://www.airarabia.com"),
    "XY": ("FLYNAS", "https://flynas.com"),
    "QR": ("QATAR AIRWAYS", "https://www.qatarairways.com"),
    "EK": ("EMIRATES", "https://www.emirates.com"),
}


def _official_airline_site(flight: dict) -> tuple[str, str] | None:
    """flight_data'dagi aviakompaniya kodi yoki to'liq nomi bo'yicha rasmiy saytni topadi.

    Moslik IATA kodi ("HY") yoki to'liq nom ("Uzbekistan Airways", har qanday
    registrda) bo'yicha tekshiriladi. Topilmasa — None qaytadi va xabarga
    "rasmiy sayt" qatori qo'shilmaydi.
    """
    code = str(flight.get("airline_code") or "").strip().upper()
    name = str(flight.get("airline") or "").strip().upper()

    # 1) IATA kodi bo'yicha (airline_code yoki airline maydonining o'zi kod bo'lsa)
    for candidate in (code, name):
        if candidate:
            entry = AIRLINE_OFFICIAL_SITES.get(candidate)
            if entry:
                return entry
    # 2) To'liq nom bo'yicha (har qanday registrda; masalan "Uzbekistan Airways")
    if name:
        for official_name, url in AIRLINE_OFFICIAL_SITES.values():
            if name == official_name or (official_name and official_name in name):
                return official_name, url
    return None


DATE_ONLY_RE = re.compile(r"^\d{4}-\d{2}-\d{2}")
GOOGLE_FLIGHTS_BASE = "https://www.google.com/travel/flights?q="


def _google_flights_url(order: dict, flight_data: dict) -> str:
    """Aynan shu reysni ochadigan Google Flights havolasi (URL-encode qilingan).

    So'rov: "flights from {origin} to {destination} on {sana}" — aviakompaniya
    nomi bo'lsa oxiriga " on {aviakompaniya}" qo'shiladi, shunda Google Flights
    aynan o'sha reysni ko'rsatadi.
    """
    origin_safe = str(order.get("origin") or "").strip()
    dest_safe = str(order.get("destination") or "").strip()

    # Sana: flight_data'dagi ISO vaqt bo'lsa uning sana qismi, aks holda buyurtma sanasi
    depart_raw = str(
        flight_data.get("departure_at")
        or flight_data.get("departure_time")
        or ""
    ).strip()
    if not DATE_ONLY_RE.match(depart_raw):
        depart_raw = str(order.get("depart_date") or "").strip()
    depart_safe = depart_raw[:10]

    f_airline = str(flight_data.get("airline") or "").strip()

    gf_query = f"flights from {origin_safe} to {dest_safe} on {depart_safe}"
    if f_airline:
        gf_query += f" on {f_airline}"
    return GOOGLE_FLIGHTS_BASE + quote(gf_query)


bot = Bot(token=settings.BOT_TOKEN)
dp = Dispatcher()
dp.include_router(bot_router)


@asynccontextmanager
async def lifespan(app: FastAPI):
    webhook_url = f"{settings.WEBHOOK_BASE_URL.rstrip('/')}/webhook" if settings.WEBHOOK_BASE_URL else ""
    if webhook_url and webhook_url.startswith("http"):
        try:
            await bot.set_webhook(webhook_url, drop_pending_updates=False)
            log.info(f"Webhook o'rnatildi: {webhook_url}")
        except Exception as e:
            log.warning(f"Webhook o'rnatishda xatolik: {e}")
    yield
    try:
        await bot.session.close()
    except Exception as e:
        log.debug(f"Bot session yopishda xatolik: {e}")


app = FastAPI(title="Saudiya Biletlar API", lifespan=lifespan)

# ==================== XAVFSIZLIK HEADERS ====================
@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    # XSS himoyasi
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    # Content Security Policy - XSS hujumlarini bloklaydi
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://telegram.org https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: https: blob:; "
        "connect-src 'self' https://api.travelpayouts.com https://cbu.uz https://autocomplete.travelpayouts.com https://www.aviasales.com; "
        "frame-ancestors 'none';"
    )
    # Cache control for sensitive pages
    path = request.url.path.lower()
    if path.startswith("/api/") or path in ("/admin", "/admin/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
    return response


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# Telegram Mini App statik fayllarni juda uzoq keshlaydi — natijada foydalanuvchi
# eski dizaynni ko'rib qoladi. HTML/JS/CSS uchun keshni butunlay o'chiramiz.
@app.middleware("http")
async def no_cache_for_static(request: Request, call_next):
    response = await call_next(request)
    path = request.url.path.lower()
    if path.endswith((".html", ".js", ".css")) or path in ("/", "/admin/", "/admin"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

# /admin -> /admin/ avtomatik redirect (StaticFiles mountdan OLDIN e'lon qilinadi)
@app.get("/admin", include_in_schema=False)
async def admin_redirect():
    return RedirectResponse(url="/admin/", status_code=307)


app.mount("/admin", StaticFiles(directory="admin_static", html=True), name="admin")


def verify_admin(x_admin_password: str = Header(default="")):
    if x_admin_password != settings.ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Noto'g'ri admin parol")
    return True


def _required_text(payload: dict, field: str, *, min_length: int = 1, max_length: int = 200) -> str:
    value = str(payload.get(field) or "").strip()
    if len(value) < min_length:
        raise HTTPException(status_code=400, detail=f"'{field}' maydoni to'ldirilmagan")
    if len(value) > max_length:
        raise HTTPException(status_code=400, detail=f"'{field}' maydoni juda uzun")
    return value


def _optional_text(payload: dict, field: str, *, max_length: int = 1000) -> str | None:
    value = str(payload.get(field) or "").strip()
    if not value:
        return None
    if len(value) > max_length:
        raise HTTPException(status_code=400, detail=f"'{field}' maydoni juda uzun")
    return value


def _telegram_user_id(payload: dict) -> int:
    try:
        user_id = int(payload.get("telegram_user_id"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Telegram foydalanuvchi ID noto'g'ri")
    if user_id <= 0:
        raise HTTPException(status_code=400, detail="Mini Appni Telegram ichidan oching")
    return user_id


def _verify_telegram_init_data(init_data: str, claimed_user_id: int) -> int:
    """Telegram Mini App initData imzosini tekshiradi va ID almashtirilishini bloklaydi."""
    if not init_data:
        raise HTTPException(status_code=401, detail="Telegram tasdiqlash ma'lumoti yo'q")
    try:
        values = dict(parse_qsl(init_data, keep_blank_values=True, strict_parsing=True))
        received_hash = values.pop("hash")
        auth_date = int(values.get("auth_date") or 0)
        user_data = json.loads(values.get("user") or "{}")
        telegram_user_id = int(user_data.get("id"))
    except (KeyError, TypeError, ValueError, json.JSONDecodeError):
        raise HTTPException(status_code=401, detail="Telegram tasdiqlash ma'lumoti noto'g'ri")

    data_check_string = "\n".join(f"{key}={values[key]}" for key in sorted(values))
    secret_key = hmac.new(b"WebAppData", settings.BOT_TOKEN.encode(), hashlib.sha256).digest()
    expected_hash = hmac.new(secret_key, data_check_string.encode(), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected_hash, received_hash):
        raise HTTPException(status_code=401, detail="Telegram imzosi noto'g'ri")

    now = int(time.time())
    if auth_date <= 0 or auth_date > now + 300 or now - auth_date > TELEGRAM_INIT_DATA_MAX_AGE:
        raise HTTPException(status_code=401, detail="Telegram sessiyasi eskirgan, Mini Appni qayta oching")
    if telegram_user_id != claimed_user_id:
        raise HTTPException(status_code=403, detail="Boshqa foydalanuvchi ma'lumotiga ruxsat yo'q")
    return telegram_user_id


def _iso_date(value: object, field: str, *, required: bool = True) -> date | None:
    raw = str(value or "").strip()
    if not raw and not required:
        return None
    try:
        return date.fromisoformat(raw)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"'{field}' sanasi noto'g'ri")


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


# ==================== TELEGRAM WEBHOOK ====================
@app.post("/webhook")
async def telegram_webhook(request: Request):
    try:
        data = await request.json()
        update = Update.model_validate(data)
        await dp.feed_update(bot, update)
    except Exception:
        log.exception("Telegram webhook xatosi")
    return {"ok": True}


# ==================== CHIPTA QIDIRUV ====================
@app.get("/api/search")
async def api_search(origin: str, destination: str, depart_date: str):
    origin_iata = tp.to_iata(origin)
    dest_iata = tp.to_iata(destination)

    # 1) Qo'lda qo'shilgan chiptalar
    try:
        manual = db.list_manual_flights(origin_iata, dest_iata, depart_date)
    except Exception:
        log.exception("manual_flights xatolik")
        manual = []

    manual_results = [{
        "origin": (f.get("origin") or origin_iata).upper(),
        "destination": (f.get("destination") or dest_iata).upper(),
        "price": f.get("price"),
        "airline": f.get("airline") or "Saudiya Biletlar",
        "flight_number": f.get("flight_number") or "SAU-001",
        "departure_at": f"{f.get('depart_date')}T{f.get('departure_time') or '09:30'}:00",
        "transfers": f.get("transfers", 0),
        "seats_available": f.get("seats_available", 10),
        "source": "manual",
        "manual_flight_id": f.get("id"),
    } for f in manual]

    # 2) Travelpayouts API — faqat ishonchli aviakompaniyalar (TRUSTED_AIRLINES)
    try:
        api_results = await tp.search_flights(origin, destination, depart_date)
        api_results = [f for f in api_results if f.get("airline") in tp.TRUSTED_AIRLINES or tp.is_trusted_offer(f)]
        api_results = [tp.enrich_partner_offer(r) for r in api_results]
        for r in api_results:
            r["source"] = "api"
    except Exception:
        log.exception("search error")
        api_results = []

    return {"results": manual_results + api_results}


# ==================== BUYURTMA YARATISH ====================
@app.post("/api/orders")
async def api_create_order(payload: dict):
    required = ["telegram_user_id", "origin", "destination", "depart_date", "flight_data", "passport"]
    for field in required:
        if field not in payload:
            raise HTTPException(status_code=400, detail=f"'{field}' maydoni yo'q")

    # flight_data string (JSON) ko'rinishida ham kelishi mumkin — parse qilinadi
    flight_data = payload.get("flight_data") or {}
    if isinstance(flight_data, str):
        try:
            flight_data = json.loads(flight_data)
        except (TypeError, ValueError):
            flight_data = {}
    if not isinstance(flight_data, dict):
        flight_data = {}
    price = payload.get("price")
    if price is None and isinstance(flight_data, dict):
        price = flight_data.get("price")

    order = db.create_order({
        "telegram_user_id": payload["telegram_user_id"],
        "username": payload.get("username"),
        "origin": str(payload["origin"]).upper(),
        "destination": str(payload["destination"]).upper(),
        "depart_date": payload["depart_date"],
        "passengers": payload.get("passengers", 1),
        "flight_data": flight_data,
        "price": price,
        "status": "new",
    })

    passport_data = payload.get("passport") or {}
    order_id = order.get("id")
    passport = db.save_passport(order_id, passport_data)

    first_n = html.escape(str(passport.get("first_name") or "-"))
    last_n = html.escape(str(passport.get("last_name") or ""))
    p_num = html.escape(str(passport.get("passport_number") or "-"))
    b_year = html.escape(str(passport.get("birth_year") or "-"))
    exp_date = html.escape(str(passport.get("expiry_date") or "-"))

    flight_info_lines = []
    if isinstance(flight_data, dict):
        airline = html.escape(str(flight_data.get("airline") or "").strip())
        flight_number = html.escape(str(flight_data.get("flight_number") or "").strip())
        departure_value = flight_data.get("departure_at") or flight_data.get("departure_time")
        departure_str = html.escape(str(departure_value).strip()) if departure_value else ""
        link_value = str(flight_data.get("link") or "").strip()

        if airline:
            airline_line = f"🛫 Aviakompaniya: {airline}"
            if flight_number:
                airline_line += f" ({flight_number})"
            flight_info_lines.append(airline_line)
        if departure_str:
            flight_info_lines.append(f"🕐 Jo'nash vaqti: {departure_str}")
        if link_value.lower().startswith(("http://", "https://")):
            safe_link = html.escape(link_value)
            flight_info_lines.append(f"🔗 Chiptani shu havoladan oling: {safe_link}")
        # Xavfsiz xarid: aynan shu reysni Google Flights'da ochadigan havola
        gf_url = _google_flights_url(order, flight_data)
        gf_href = html.escape(gf_url, quote=True)
        flight_info_lines.append(
            f"✅ <b>Xavfsiz xarid:</b> <a href=\"{gf_href}\">O'sha reysni ochish ➔</a>"
        )
        # Aviakompaniya rasmiy sayti (kod yoki to'liq nom bo'yicha) — XSS himoyasi bilan
        official = _official_airline_site(flight_data)
        if official:
            official_name, official_url = official
            flight_info_lines.append(
                f"🏢 <b>Rasmiy sayt:</b> {html.escape(official_name)} — {html.escape(official_url)}"
            )

    flight_info_block = ""
    if flight_info_lines:
        flight_info_block = "\n".join(flight_info_lines) + "\n\n"

    text = (
        f"🆕 <b>Yangi buyurtma #{order_id}</b>\n\n"
        f"👤 {first_n} {last_n}\n"
        f"🛂 Passport: {p_num}\n"
        f"📅 Tug'ilgan yil: {b_year}\n"
        f"⏳ Amal qilish: {exp_date}\n\n"
        f"✈️ {order.get('origin')} ➔ {order.get('destination')}\n"
        f"🗓 {order.get('depart_date')} | 👥 {order.get('passengers', 1)} yo'lovchi\n"
        f"💵 <b>${order.get('price', '-')} USD</b>\n\n"
        f"{flight_info_block}"
        f"👇 Pastdagi tugmalar orqali 1 bosishda boshqaring:"
    )
    try:
        await bot.send_message(
            settings.ADMIN_CHAT_ID,
            text,
            parse_mode="HTML",
            reply_markup=admin_order_keyboard(order_id),
        )
    except Exception as e:
        log.warning(f"Admin guruhiga xabar yuborilmadi: {e}")

    return {"order_id": order_id}


# ==================== TO'LOV CHEKI ====================
@app.post("/api/orders/{order_id}/payment")
async def api_upload_payment(order_id: int, file: UploadFile = File(...)):
    """To'lov cheki yuklash - fayl turi va hajim xavfsizligi bilan."""
    order = db.get_order(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Buyurtma topilmadi")

    content = await file.read()
    filename = html.escape(str(file.filename or "receipt.jpg"))
    
    # Xavfsizlik: fayl turi va hajmini tekshirish
    try:
        _validate_payment_file(filename, content, file.content_type or "")
    except HTTPException:
        raise
    except Exception as e:
        log.warning(f"Fayl tekshirishda xatolik: {e}")
        raise HTTPException(status_code=400, detail="Fayl noto'g'ri")
    
    # Xavfsiz filename - path traversal himoyasi
    safe_filename = os.path.basename(filename)
    storage_path = f"{order_id}_{safe_filename}"
    
    url = db.upload_file("payments", storage_path, content, file.content_type or "image/jpeg")
    db.update_order(order_id, {"payment_screenshot_url": url, "status": "awaiting_confirmation"})

    try:
        photo = BufferedInputFile(content, filename=safe_filename)
        safe_origin = html.escape(str(order.get('origin') or '').upper())
        safe_dest = html.escape(str(order.get('destination') or '').upper())
        safe_date = html.escape(str(order.get('depart_date') or '-'))
        safe_price = html.escape(str(order.get('price', '-')))
        await bot.send_photo(
            settings.ADMIN_CHAT_ID,
            photo,
            caption=(
                f"💳 <b>To'lov cheki — buyurtma #{order_id}</b>\n\n"
                f"✈️ {safe_origin} ➔ {safe_dest}\n"
                f"🗓 {safe_date} | 💵 <b>${safe_price}</b>\n\n"
                f"👇 1 bosishda tasdiqlang yoki rad eting:"
            ),
            parse_mode="HTML",
            reply_markup=admin_order_keyboard(order_id),
        )
    except Exception as e:
        log.warning(f"Admin guruhiga to'lov cheki yuborilmadi: {e}")

    return {"ok": True, "url": url}


# ==================== MIJOZNING BUYURTMALARI TARIXI ====================
@app.get("/api/my-orders")
async def api_my_orders(telegram_user_id: int):
    try:
        orders = db.get_orders_by_user(telegram_user_id, limit=20)
        return {"orders": orders}
    except Exception:
        log.exception("my-orders xatolik")
        return {"orders": []}


# ==================== VIZA ARIZALARI ====================
@app.post("/api/visa-applications")
async def api_create_visa_application(
    payload: dict,
    x_telegram_init_data: str = Header(default=""),
):
    user_id = _telegram_user_id(payload)
    _verify_telegram_init_data(x_telegram_init_data, user_id)
    visa_type = str(payload.get("visa_type") or "").strip()
    if visa_type not in VISA_TYPES:
        raise HTTPException(status_code=400, detail="Viza turi noto'g'ri")

    first_name = _required_text(payload, "first_name", min_length=2, max_length=80).upper()
    last_name = _required_text(payload, "last_name", min_length=2, max_length=80).upper()
    phone = _required_text(payload, "phone", min_length=7, max_length=30)
    passport_number = _required_text(payload, "passport_number", min_length=5, max_length=20).upper()
    if not PASSPORT_RE.fullmatch(passport_number):
        raise HTTPException(status_code=400, detail="Pasport raqami faqat lotin harflari va raqamlardan iborat bo'lsin")

    birth_date = _iso_date(payload.get("birth_date"), "birth_date")
    travel_date = _iso_date(payload.get("travel_date"), "travel_date", required=False)
    if birth_date >= date.today():
        raise HTTPException(status_code=400, detail="Tug'ilgan sana noto'g'ri")
    if travel_date and travel_date < date.today():
        raise HTTPException(status_code=400, detail="Safar sanasi o'tgan bo'lishi mumkin emas")

    application = db.create_visa_application({
        "telegram_user_id": user_id,
        "username": _optional_text(payload, "username", max_length=64),
        "visa_type": visa_type,
        "first_name": first_name,
        "last_name": last_name,
        "phone": phone,
        "passport_number": passport_number,
        "birth_date": birth_date.isoformat(),
        "travel_date": travel_date.isoformat() if travel_date else None,
        "notes": _optional_text(payload, "notes", max_length=1000),
        "status": "new",
    })
    application_id = application.get("id")

    safe_name = html.escape(f"{first_name} {last_name}")
    safe_phone = html.escape(phone)
    safe_passport = html.escape(passport_number)
    visa_label = html.escape(VISA_TYPES[visa_type])
    text = (
        f"📑 <b>Yangi viza arizasi #{application_id or '-'}</b>\n\n"
        f"🕋 Turi: <b>{visa_label}</b>\n"
        f"👤 Arizachi: {safe_name}\n"
        f"📞 Telefon: {safe_phone}\n"
        f"🛂 Pasport: <code>{safe_passport}</code>\n"
        f"📅 Rejalashtirilgan safar: {travel_date.isoformat() if travel_date else '-'}\n"
        f"💬 Telegram ID: <code>{user_id}</code>\n\n"
        "Arizani Admin paneldagi «Viza arizalari» bo'limida boshqaring."
    )
    try:
        await bot.send_message(settings.ADMIN_CHAT_ID, text, parse_mode="HTML")
    except Exception as e:
        log.warning(f"Yangi viza arizasi haqida xabar yuborilmadi: {e}")

    return {"application_id": application_id, "application": application}


@app.get("/api/visa-applications")
async def api_my_visa_applications(
    telegram_user_id: int,
    x_telegram_init_data: str = Header(default=""),
):
    if telegram_user_id <= 0:
        raise HTTPException(status_code=400, detail="Telegram foydalanuvchi ID noto'g'ri")
    _verify_telegram_init_data(x_telegram_init_data, telegram_user_id)
    return {"applications": db.get_visa_applications_by_user(telegram_user_id)}


# ==================== NARX TUSHGANDA OBUNA ====================
@app.post("/api/price-alerts")
async def api_create_price_alert(
    payload: dict,
    x_telegram_init_data: str = Header(default=""),
):
    user_id = _telegram_user_id(payload)
    _verify_telegram_init_data(x_telegram_init_data, user_id)
    origin = tp.to_iata(_required_text(payload, "origin", max_length=80)).upper()
    destination = tp.to_iata(_required_text(payload, "destination", max_length=80)).upper()
    if not IATA_RE.fullmatch(origin) or not IATA_RE.fullmatch(destination):
        raise HTTPException(status_code=400, detail="Aeroport kodi noto'g'ri")
    if origin == destination:
        raise HTTPException(status_code=400, detail="Jo'nash va borish aeroporti bir xil bo'lishi mumkin emas")

    date_from = _iso_date(payload.get("date_from"), "date_from")
    date_to = _iso_date(payload.get("date_to"), "date_to")
    if date_from < date.today():
        raise HTTPException(status_code=400, detail="Obuna boshlanish sanasi o'tgan bo'lishi mumkin emas")
    if date_to < date_from:
        raise HTTPException(status_code=400, detail="Obuna tugash sanasi boshlanish sanasidan oldin")
    if (date_to - date_from).days >= MAX_ALERT_RANGE_DAYS:
        raise HTTPException(status_code=400, detail=f"Obuna oralig'i ko'pi bilan {MAX_ALERT_RANGE_DAYS} kun")

    try:
        target_price = round(float(payload.get("target_price")), 2)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Maqsadli narx noto'g'ri")
    if target_price <= 0 or target_price > 100000:
        raise HTTPException(status_code=400, detail="Maqsadli narx musbat son bo'lishi kerak")

    alert = db.create_price_alert({
        "telegram_user_id": user_id,
        "username": _optional_text(payload, "username", max_length=64),
        "origin": origin,
        "destination": destination,
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "target_price": target_price,
        "is_active": True,
    })
    return {"alert_id": alert.get("id"), "alert": alert}


@app.get("/api/price-alerts")
async def api_my_price_alerts(
    telegram_user_id: int,
    active_only: bool = False,
    x_telegram_init_data: str = Header(default=""),
):
    if telegram_user_id <= 0:
        raise HTTPException(status_code=400, detail="Telegram foydalanuvchi ID noto'g'ri")
    _verify_telegram_init_data(x_telegram_init_data, telegram_user_id)
    return {"alerts": db.get_price_alerts_by_user(telegram_user_id, active_only=active_only)}


@app.delete("/api/price-alerts/{alert_id}")
async def api_cancel_price_alert(
    alert_id: int,
    telegram_user_id: int,
    x_telegram_init_data: str = Header(default=""),
):
    _verify_telegram_init_data(x_telegram_init_data, telegram_user_id)
    alert = db.get_price_alert(alert_id)
    if not alert or int(alert.get("telegram_user_id") or 0) != telegram_user_id:
        raise HTTPException(status_code=404, detail="Obuna topilmadi")
    db.update_price_alert(alert_id, {"is_active": False})
    return {"ok": True, "alert_id": alert_id}


async def _best_price_for_alert(alert: dict) -> dict | None:
    """Faqat API yoki admin kiritgan haqiqiy narxlardan eng arzonini topadi.

    Taqvimdagi deterministik `estimate` qiymatlar obuna xabarini ishga tushirmaydi.
    """
    start = _iso_date(alert.get("date_from"), "date_from")
    end = _iso_date(alert.get("date_to"), "date_to")
    days = min(MAX_ALERT_RANGE_DAYS, (end - start).days + 1)
    origin = str(alert.get("origin") or "").upper()
    destination = str(alert.get("destination") or "").upper()

    candidates: list[dict] = []
    calendar = await tp.get_calendar_prices(origin, destination, start.isoformat(), days)
    for item in calendar:
        if item.get("source") not in {"api", "manual"} or item.get("price") is None:
            continue
        candidates.append({
            "price": float(item["price"]),
            "date": item.get("date"),
            "source": item.get("source"),
            "airline": item.get("airline") or "",
        })

    try:
        manual = db.list_manual_flights(origin, destination, None)
    except Exception:
        manual = []
    for item in manual:
        depart = _iso_date(item.get("depart_date"), "depart_date", required=False)
        if not depart or depart < start or depart > end or item.get("price") is None:
            continue
        try:
            price = float(item["price"])
        except (TypeError, ValueError):
            continue
        candidates.append({
            "price": price,
            "date": depart.isoformat(),
            "source": "manual",
            "airline": item.get("airline") or "Saudiya Biletlar",
        })

    return min(candidates, key=lambda item: item["price"]) if candidates else None


@app.post("/api/cron/price-alerts")
@app.get("/api/cron/price-alerts")
async def api_check_price_alerts(secret: str = "", limit: int = 100):
    """Faol obunalarni tekshiradi va shart bajarilganda Telegram xabari yuboradi."""
    if secret != settings.CRON_SECRET:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    alerts = db.list_price_alerts(active_only=True, limit=max(1, min(int(limit or 100), 200)))
    checked = matched = notified = 0
    errors: list[dict] = []
    now_iso = _utc_now_iso()

    for alert in alerts:
        alert_id = alert.get("id")
        checked += 1
        try:
            best = await _best_price_for_alert(alert)
            update_data: dict = {"last_checked_at": now_iso}
            if best:
                update_data["last_price"] = round(float(best["price"]), 2)

            if not best or float(best["price"]) > float(alert.get("target_price") or 0):
                db.update_price_alert(alert_id, update_data)
                continue

            matched += 1
            origin = html.escape(str(alert.get("origin") or ""))
            destination = html.escape(str(alert.get("destination") or ""))
            depart_date = html.escape(str(best.get("date") or ""))
            airline = html.escape(str(best.get("airline") or "Aviakompaniya"))
            price = round(float(best["price"]), 2)
            target = round(float(alert.get("target_price") or 0), 2)
            message = (
                "🔔 <b>Narx tushdi!</b>\n\n"
                f"✈️ <b>{origin} ➔ {destination}</b>\n"
                f"📅 Sana: <b>{depart_date}</b>\n"
                f"🛫 {airline}\n"
                f"💵 Hozirgi narx: <b>${price:g}</b>\n"
                f"🎯 Siz belgilagan narx: ${target:g}\n\n"
                "Taklifni o'tkazib yubormaslik uchun Mini Appga kirib band qiling."
            )
            await bot.send_message(int(alert["telegram_user_id"]), message, parse_mode="HTML")
            update_data.update({"last_notified_at": now_iso, "is_active": False})
            db.update_price_alert(alert_id, update_data)
            notified += 1
        except Exception as e:
            log.exception(f"Narx obunasini tekshirishda xatolik #{alert_id}")
            errors.append({"alert_id": alert_id, "error": str(e)[:200]})

    return {
        "checked": checked,
        "matched": matched,
        "notified": notified,
        "errors": errors,
    }


# ==================== MARKAZIY BANK (CBU) JONLI KURSI ====================
CBU_USD_URL = "https://cbu.uz/uz/arkhiv-kursov-valyut/json/USD/"
_cbu_cache: dict = {"rate": None, "date": None, "diff": None, "ts": 0.0}
CBU_CACHE_TTL = 30 * 60  # 30 daqiqa
CBU_FALLBACK_RATE = 12850.0


async def get_cbu_usd_rate() -> dict:
    """Markaziy Bankdan (cbu.uz) USD/UZS jonli kursini oladi (30 daqiqa kesh)."""
    now = time.time()
    if _cbu_cache["rate"] and (now - _cbu_cache["ts"]) < CBU_CACHE_TTL:
        return {
            "rate": _cbu_cache["rate"],
            "date": _cbu_cache["date"],
            "diff": _cbu_cache["diff"],
            "source": "cbu.uz (kesh)",
        }

    try:
        async with httpx.AsyncClient(timeout=10) as client:
            r = await client.get(CBU_USD_URL)
            r.raise_for_status()
            data = r.json()
            item = data[0] if isinstance(data, list) and data else data
            rate = float(str(item.get("Rate")).replace(",", "."))
            _cbu_cache.update({
                "rate": rate,
                "date": item.get("Date"),
                "diff": item.get("Diff"),
                "ts": now,
            })
            return {"rate": rate, "date": item.get("Date"), "diff": item.get("Diff"), "source": "cbu.uz"}
    except Exception as e:
        log.warning(f"CBU kursini olishda xatolik: {e}")
        return {
            "rate": _cbu_cache["rate"] or CBU_FALLBACK_RATE,
            "date": _cbu_cache["date"],
            "diff": _cbu_cache["diff"],
            "source": "zaxira",
        }


@app.get("/api/cbu-rate")
async def api_cbu_rate():
    """Markaziy Bankning jonli USD kursi (Mini App va admin panel uchun)."""
    return await get_cbu_usd_rate()


# ==================== 🔥 AVTO NARX TAVSIYALARI (TOP DEALS) ====================
_deals_cache: dict = {"data": None, "ts": 0.0}
DEALS_CACHE_TTL = 30 * 60  # 30 daqiqa


@app.get("/api/top-deals")
async def api_top_deals(limit: int = 8, refresh: bool = False):
    """Mini App ochilishida avtomatik ko'rsatiladigan eng arzon takliflar.

    O'zbekistonning 11 ta aeroportidan Jidda/Madinaga, faqat yaqin 3–35 kun
    ichidagi reyslar. Narxlar 30 daqiqa keshlanadi (avtomatik yangilanadi).
    """
    limit = max(1, min(int(limit or 8), 11))
    now = time.time()

    cached = _deals_cache.get("data")
    if cached and not refresh and (now - float(_deals_cache.get("ts") or 0)) < DEALS_CACHE_TTL:
        deals = cached
    else:
        try:
            tickets = await tp.get_daily_cheapest()
        except Exception:
            log.exception("Top-deals narxlarini olishda xatolik")
            tickets = []

        valid = tp.filter_offers_by_window([t for t in (tickets or []) if t.get("value") is not None])
        valid = tp.top_up_missing_cities(valid)
        picked = tp.pick_mixed_offers(valid, limit=11)

        deals = []
        for t in picked:
            try:
                price = round(float(t.get("value")), 2)
            except (TypeError, ValueError):
                continue
            origin = str(t.get("origin") or "").upper()
            dest = str(t.get("destination") or "").upper()
            deals.append({
                "origin": origin,
                "origin_name": t.get("origin_name") or tp.city_name(origin),
                "destination": dest,
                "destination_name": t.get("destination_name") or tp.city_name(dest),
                "price": price,
                "depart_date": t.get("depart_date"),
                "depart_date_label": t.get("depart_date_label") or tp.format_date_uz(t.get("depart_date")),
                "days_left": t.get("days_left"),
                "source": t.get("source") or "api",
            })

        deals.sort(key=lambda d: d["price"])
        for i, d in enumerate(deals):
            d["is_cheapest"] = i == 0
        _deals_cache.update({"data": deals, "ts": now})

    rate_info = await get_cbu_usd_rate()
    return {
        "deals": deals[:limit],
        "rate": rate_info.get("rate") or CBU_FALLBACK_RATE,
        "window": {"min_days": tp.MIN_DAYS_AHEAD, "max_days": tp.MAX_DAYS_AHEAD},
        "updated_at": datetime.now().strftime("%H:%M"),
    }


# ==================== ARZON NARXLAR TAQVIMI ====================
@app.get("/api/calendar")
async def api_calendar(
    origin: str = "TAS",
    destination: str = "JED",
    start_date: str | None = None,
    days: int = 30,
):
    """Har bir kun bo'yicha eng arzon narxlar (gorizontal taqvim uchun)."""
    origin_iata = tp.to_iata(origin)
    dest_iata = tp.to_iata(destination)

    try:
        calendar = await tp.get_calendar_prices(origin_iata, dest_iata, start_date, days)
    except Exception:
        log.exception("Taqvim narxlarini olishda xatolik")
        calendar = []

    # Qo'lda qo'shilgan chiptalar taqvimdagi narxlardan arzon bo'lsa — ularni ustun qo'yamiz
    try:
        manual = db.list_manual_flights(origin_iata, dest_iata, None)
    except Exception:
        manual = []

    manual_by_date: dict[str, float] = {}
    for f in manual:
        d = str(f.get("depart_date") or "")[:10]
        try:
            price = float(f.get("price"))
        except (TypeError, ValueError):
            continue
        if d and (d not in manual_by_date or price < manual_by_date[d]):
            manual_by_date[d] = price

    for day in calendar:
        m_price = manual_by_date.get(day["date"])
        if m_price is not None and (day.get("price") is None or m_price < float(day["price"])):
            day["price"] = m_price
            day["source"] = "manual"
            day["airline"] = "Saudiya Biletlar"

    prices = [float(d["price"]) for d in calendar if d.get("price") is not None]
    cheapest_price = min(prices) if prices else None
    cheapest_date = None
    if cheapest_price is not None:
        for d in calendar:
            if d.get("price") is not None and float(d["price"]) == cheapest_price:
                cheapest_date = d["date"]
                break
        for d in calendar:
            d["is_cheapest"] = d.get("date") == cheapest_date

    return {
        "origin": origin_iata,
        "destination": dest_iata,
        "origin_name": tp.city_name(origin_iata),
        "destination_name": tp.city_name(dest_iata),
        "cheapest_price": cheapest_price,
        "cheapest_date": cheapest_date,
        "days": calendar,
    }


# ==================== KANALGA CHIROYLI AVTO-POST ====================
@app.post("/api/cron/daily-post")
@app.get("/api/cron/daily-post")
async def api_daily_post(
    secret: str = "",
    limit: int = 11,
    min_days: int = tp.MIN_DAYS_AHEAD,
    max_days: int = tp.MAX_DAYS_AHEAD,
):
    """O'zbekistonning 11 ta xalqaro aeroportidan Jidda va Madinaga aralash reyslar posti.

    Faqat yaqin `min_days`–`max_days` (sukut bo'yicha 3–35) kun ichidagi sanalar chiqadi —
    uzoq dekabr/yanvar sanalari postga tushmaydi.
    """
    if secret != settings.CRON_SECRET:
        raise HTTPException(status_code=403, detail="Ruxsat yo'q")

    # Cron parametrlari ham qat'iy 3–35 kunlik oynadan chiqmasin.
    # Operator ichki oynani qisqartirishi mumkin, lekin uzoq sanalarni yoqolmaydi.
    try:
        requested_min = int(min_days)
        requested_max = int(max_days)
    except (TypeError, ValueError):
        requested_min, requested_max = tp.MIN_DAYS_AHEAD, tp.MAX_DAYS_AHEAD
    min_days = max(tp.MIN_DAYS_AHEAD, min(requested_min, tp.MAX_DAYS_AHEAD))
    max_days = max(min_days, min(requested_max, tp.MAX_DAYS_AHEAD))

    try:
        tickets = await tp.get_daily_cheapest(min_days=min_days, max_days=max_days)
    except Exception:
        log.exception("Kunlik narxlarni olishda xatolik")
        tickets = []

    # 1) Faqat narxi bor va sanasi 3–35 kun oynasiga tushadigan takliflar
    valid_tickets = tp.filter_offers_by_window(
        [tp.enrich_partner_offer(t) for t in (tickets or []) if t.get("value") is not None],
        min_days=min_days,
        max_days=max_days,
    )

    used_fallback = not valid_tickets

    # 2) API'dan tushmagan shaharlar zaxira (taxminiy) narxlar bilan to'ldiriladi,
    #    shunda kanal bo'sh qolmaydi va 11 ta aeroport ham qatnashadi
    valid_tickets = tp.top_up_missing_cities(
        valid_tickets,
        min_days=min_days,
        max_days=max_days,
    )

    # Turfa xil aralash reyslar: bitta shahar (aeroport) takrorlanmaydi
    selected = tp.pick_mixed_offers(valid_tickets, limit=max(1, min(int(limit or 11), 11)))
    if not selected:
        return {"posted": 0}

    rate_info = await get_cbu_usd_rate()
    uzs_rate = float(rate_info.get("rate") or CBU_FALLBACK_RATE)

    today_str = datetime.now().strftime("%d.%m.%Y")
    text = (
        "🕋 <b>SAUDIYA BILETLAR | BUGUNGI ENG ARZON REYSLAR</b> ✈️\n"
        f"📆 <i>{today_str}</i> — O'zbekistonning barcha aeroportlaridan Jidda va Madinaga\n"
        f"🗓 <i>Faqat yaqin {min_days}–{max_days} kun ichidagi reyslar:</i>\n\n"
    )

    for t in selected:
        origin = str(t.get("origin") or "").upper()
        dest = str(t.get("destination") or "").upper()
        val = int(float(t.get("value") or 380))
        val_uzs = f"{int(val * uzs_rate):,}".replace(",", " ")
        origin_name = t.get("origin_name") or tp.city_name(origin)
        dest_name = t.get("destination_name") or tp.city_name(dest)
        dest_icon = "🕋" if dest == "MED" else "🌅"
        date_label = t.get("depart_date_label") or tp.format_date_uz(t.get("depart_date"))
        days_left = t.get("days_left")
        if days_left is None:
            days_left = tp.days_until(t.get("depart_date"))
        days_note = f" — {days_left} kundan keyin" if days_left is not None else ""
        airline_label = str(t.get("airline_label") or t.get("airline") or "").strip()
        airline_line = f"   ✈️ {airline_label}\n" if airline_label else ""
        text += (
            f"{dest_icon} <b>{origin_name} ({origin}) ➔ {dest_name} ({dest})</b>\n"
            f"   📅 Sana: <code>{date_label}</code>{days_note}\n"
            f"   💵 Narxi: <b>${val}</b> (~{val_uzs} so'm)\n"
            f"{airline_line}"
            f"   🧳 Bagaj: 30 kg + 7 kg | 🍽 Issiq taom bepul\n"
            f"   ──────────────\n"
        )

    text += (
        f"\n💱 Markaziy Bank kursi: <b>1$ = {int(uzs_rate):,}</b> so'm\n".replace(",", " ")
        + "⚡️ <i>Joylar soni cheklangan! Chipta band qilish uchun botga kiring:</i>\n"
        "👉 @Saudiya_Biletlarbot\n"
        f"👤 Savollar uchun: @{settings.ADMIN_USERNAME}"
    )

    try:
        await bot.send_message(settings.CHANNEL_ID, text, parse_mode="HTML")
        return {
            "posted": len(selected),
            "fallback": used_fallback,
            "window": {"min_days": min_days, "max_days": max_days},
            "cities": [str(t.get("origin") or "").upper() for t in selected],
            "dates": [t.get("depart_date") for t in selected],
        }
    except Exception as e:
        log.exception("Kanalga post yuborishda xatolik")
        raise HTTPException(status_code=500, detail=f"Kanalga xabar yuborishda xatolik: {e}")


# ==================== ADMIN PANEL ====================
@app.get("/api/admin/orders", dependencies=[Depends(verify_admin)])
async def admin_list_orders(status: str | None = None):
    orders = db.get_orders_with_passport(status=status)
    return {"orders": orders}


@app.get("/api/admin/visa-applications", dependencies=[Depends(verify_admin)])
async def admin_list_visa_applications(status: str | None = None):
    if status and status not in VISA_STATUSES:
        raise HTTPException(status_code=400, detail="Viza arizasi holati noto'g'ri")
    return {"applications": db.list_visa_applications(status=status)}


@app.patch("/api/admin/visa-applications/{application_id}", dependencies=[Depends(verify_admin)])
async def admin_update_visa_application(application_id: int, payload: dict):
    application = db.get_visa_application(application_id)
    if not application:
        raise HTTPException(status_code=404, detail="Viza arizasi topilmadi")

    status = str(payload.get("status") or "").strip()
    if status not in VISA_STATUSES:
        raise HTTPException(status_code=400, detail="Viza arizasi holati noto'g'ri")
    admin_note = _optional_text(payload, "admin_note", max_length=1000)
    updated = db.update_visa_application(application_id, {
        "status": status,
        "admin_note": admin_note,
    })

    status_labels = {
        "new": "Yangi",
        "processing": "Ko'rib chiqilmoqda",
        "approved": "Tasdiqlandi",
        "rejected": "Rad etildi",
    }
    try:
        note_line = f"\n📝 Izoh: {html.escape(admin_note)}" if admin_note else ""
        await bot.send_message(
            int(application["telegram_user_id"]),
            (
                f"📑 <b>Viza arizasi #{application_id} yangilandi</b>\n\n"
                f"Holati: <b>{html.escape(status_labels[status])}</b>"
                f"{note_line}"
            ),
            parse_mode="HTML",
        )
    except Exception as e:
        log.warning(f"Viza holati haqida mijozga xabar yuborilmadi: {e}")

    return {"application": updated}


@app.delete("/api/admin/visa-applications/{application_id}", dependencies=[Depends(verify_admin)])
async def admin_delete_visa_application(application_id: int):
    db.delete_visa_application(application_id)
    return {"ok": True, "deleted_id": application_id}


@app.get("/api/admin/price-alerts", dependencies=[Depends(verify_admin)])
async def admin_list_price_alerts(active_only: bool = False):
    return {"alerts": db.list_price_alerts(active_only=active_only)}


@app.delete("/api/admin/price-alerts/{alert_id}", dependencies=[Depends(verify_admin)])
async def admin_delete_price_alert(alert_id: int):
    db.delete_price_alert(alert_id)
    return {"ok": True, "deleted_id": alert_id}


def _generate_excel_bytes(orders: list[dict], as_csv: bool = False) -> bytes:
    """Admin uchun buyurtmalarni Excel (xlsx) yoki CSV fayl sifatida generatsiya qiladi."""
    if HAS_OPENPYXL and not as_csv:
        wb = Workbook()
        ws = wb.active
        ws.title = "Buyurtmalar"

        headers = [
            "ID",
            "Ism",
            "Familiya",
            "Pasport Raqami",
            "Tug'ilgan yil",
            "Pasport Muddati",
            "Yo'nalish",
            "Sana",
            "Yo'lovchilar",
            "Narx ($)",
            "Holati",
            "Telegram ID",
            "Username",
            "To'lov Cheki URL",
            "Yaratilgan sana",
        ]

        header_fill = PatternFill(start_color="0F5132", end_color="0F5132", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )
        center = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        STATUS_UZ = {
            "new": "Yangi",
            "awaiting_confirmation": "Tasdiq kutilmoqda",
            "confirmed": "Tasdiqlangan",
            "rejected": "Rad etilgan",
        }

        for row_idx, order in enumerate(orders, 2):
            p_raw = order.get("passports")
            if isinstance(p_raw, list) and len(p_raw) > 0 and isinstance(p_raw[0], dict):
                passport = p_raw[0]
            elif isinstance(p_raw, dict):
                passport = p_raw
            else:
                passport = {}

            route = f"{(order.get('origin') or '').upper()} -> {(order.get('destination') or '').upper()}"

            row = [
                order.get("id"),
                passport.get("first_name") or "",
                passport.get("last_name") or "",
                passport.get("passport_number") or "",
                passport.get("birth_year") or "",
                passport.get("expiry_date") or "",
                route,
                order.get("depart_date") or "",
                order.get("passengers") or 1,
                order.get("price") or "",
                STATUS_UZ.get(order.get("status"), order.get("status") or ""),
                order.get("telegram_user_id") or "",
                order.get("username") or "",
                order.get("payment_screenshot_url") or "",
                str(order.get("created_at") or "")[:19],
            ]

            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = left_align if col_idx in (2, 3, 4, 13, 14) else center
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        # Auto width
        widths = [6, 14, 14, 16, 10, 14, 16, 12, 10, 10, 18, 14, 14, 28, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
    else:
        # Fallback: CSV (Excel ham ochadi)
        import csv

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "ID", "Ism", "Familiya", "Pasport", "Tugilgan yil", "Muddat",
            "Yonalish", "Sana", "Yolovchilar", "Narx", "Holati", "Telegram ID", "Username", "Chek URL"
        ])
        for order in orders:
            p_raw = order.get("passports")
            if isinstance(p_raw, list) and p_raw and isinstance(p_raw[0], dict):
                passport = p_raw[0]
            elif isinstance(p_raw, dict):
                passport = p_raw
            else:
                passport = {}
            writer.writerow([
                order.get("id"),
                passport.get("first_name"),
                passport.get("last_name"),
                passport.get("passport_number"),
                passport.get("birth_year"),
                passport.get("expiry_date"),
                f"{order.get('origin')}->{order.get('destination')}",
                order.get("depart_date"),
                order.get("passengers"),
                order.get("price"),
                order.get("status"),
                order.get("telegram_user_id"),
                order.get("username"),
                order.get("payment_screenshot_url"),
            ])
        return buf.getvalue().encode("utf-8-sig")


@app.get("/api/admin/orders/export", dependencies=[Depends(verify_admin)])
async def admin_export_orders(status: str | None = None, format: str = "csv"):
    """Buyurtmalarni Excel (xlsx) yoki CSV formatida eksport qilish.

    format=csv  -> Excel ochadigan CSV (UTF-8 BOM bilan)
    format=xlsx -> chiroyli formatlangan Excel fayl (openpyxl bo'lsa)
    """
    orders = db.get_orders_with_passport(status=status if status else None)

    fmt = (format or "csv").strip().lower()
    as_csv = fmt == "csv" or not HAS_OPENPYXL
    data_bytes = _generate_excel_bytes(orders, as_csv=as_csv)

    ext = "csv" if as_csv else "xlsx"
    filename = f"buyurtmalar_{datetime.now().strftime('%Y-%m-%d')}.{ext}"
    media_type = (
        "text/csv; charset=utf-8"
        if as_csv
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return StreamingResponse(
        io.BytesIO(data_bytes),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.post("/api/admin/orders/{order_id}/confirm", dependencies=[Depends(verify_admin)])
async def admin_confirm_order(order_id: int):
    result = await confirm_order_action(bot, order_id)
    if not result["ok"]:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@app.post("/api/admin/orders/{order_id}/reject", dependencies=[Depends(verify_admin)])
async def admin_reject_order(order_id: int, payload: dict):
    reason = payload.get("reason", "To'lov tasdiqlanmadi")
    result = await reject_order_action(bot, order_id, reason)
    if not result["ok"]:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@app.delete("/api/admin/orders/rejected", dependencies=[Depends(verify_admin)])
@app.post("/api/admin/orders/clear-rejected", dependencies=[Depends(verify_admin)])
async def admin_clear_rejected_orders():
    """[🗑 Rad etilganlarni tozalash] — barcha rad etilgan buyurtmalarni o'chiradi."""
    try:
        deleted = db.delete_orders_by_status("rejected")
        return {"ok": True, "deleted": deleted}
    except Exception as e:
        log.exception("Rad etilgan buyurtmalarni tozalashda xatolik")
        raise HTTPException(status_code=500, detail=f"Tozalashda xatolik: {e}")


@app.delete("/api/admin/orders/{order_id}", dependencies=[Depends(verify_admin)])
async def admin_delete_order(order_id: int):
    """[🗑 O'chirish] - buyurtmani to'liq o'chirish."""
    try:
        db.delete_order(order_id)
        return {"ok": True, "deleted_id": order_id}
    except Exception as e:
        log.exception(f"Buyurtmani o'chirishda xatolik #{order_id}")
        raise HTTPException(status_code=500, detail=f"O'chirishda xatolik: {e}")


@app.delete("/api/admin/orders", dependencies=[Depends(verify_admin)])
async def admin_delete_all_orders(payload: dict = None):
    """[🗑 Barcha Buyurtmalarni O'chirish] - barcha buyurtmalarni o'chirish."""
    try:
        deleted = db.delete_all_orders()
        return {"ok": True, "deleted": deleted}
    except Exception as e:
        log.exception("Barcha buyurtmalarni o'chirishda xatolik")
        raise HTTPException(status_code=500, detail=f"Tozalashda xatolik: {e}")


@app.get("/api/admin/flights", dependencies=[Depends(verify_admin)])
async def admin_list_flights():
    return {"flights": db.list_all_manual_flights()}


@app.post("/api/admin/flights", dependencies=[Depends(verify_admin)])
async def admin_create_flight(payload: dict):
    required = ["origin", "destination", "depart_date", "price"]
    for field in required:
        if field not in payload or payload[field] is None or str(payload[field]).strip() == "":
            raise HTTPException(status_code=400, detail=f"'{field}' maydoni to'ldirilmagan")

    origin_code = tp.to_iata(str(payload["origin"])).lower()
    dest_code = tp.to_iata(str(payload["destination"])).lower()

    try:
        price = float(payload["price"])
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Narx son bo'lishi kerak")

    seats = payload.get("seats_available")
    if seats is not None:
        try:
            seats = int(seats)
        except (ValueError, TypeError):
            seats = None

    flight = db.create_manual_flight({
        "origin": origin_code,
        "destination": dest_code,
        "depart_date": str(payload["depart_date"]).strip(),
        "departure_time": payload.get("departure_time") or None,
        "price": price,
        "airline": payload.get("airline") or "Saudiya Biletlar",
        "flight_number": payload.get("flight_number") or None,
        "transfers": int(payload.get("transfers", 0) or 0),
        "seats_available": seats,
        "is_active": True,
    })
    return {"flight": flight}


@app.delete("/api/admin/flights/{flight_id}", dependencies=[Depends(verify_admin)])
async def admin_delete_flight(flight_id: int):
    db.delete_manual_flight(flight_id)
    return {"ok": True}


@app.post("/api/admin/login")
async def admin_login(request: Request, payload: dict):
    """Admin tizimga kirish - brute-force va timing-attack himoyasi bilan."""
    client_ip = _get_client_ip(request)
    
    # Brute-force tekshirish
    if _check_brute_force(client_ip):
        raise HTTPException(
            status_code=429,
            detail="Juda ko'p urinish. Iltimos, 15 daqiqadan so'ng qayta urinib ko'ring."
        )
    
    provided_password = payload.get("password", "")
    
    # Timing-attack himoyasi: secrets.compare_digest ishlatish
    # Bu har doim bir xil vaqt davom etadi, tajovuzkor parolni tong qila olmaydi
    if not provided_password:
        _record_failed_attempt(client_ip)
        raise HTTPException(status_code=401, detail="Parol kiritilmagan")
    
    # Qat'iy solishtirish - timing attack himoyasi
    is_correct = secrets.compare_digest(provided_password, settings.ADMIN_PASSWORD)
    
    if not is_correct:
        _record_failed_attempt(client_ip)
        # Xato xabari har doim bir xil uzunlikda bo'lishi uchun
        raise HTTPException(status_code=401, detail="Noto'g'ri parol")
    
    # Muvaffaqatli kirish - brute-force cheklovini olib tashlash
    _clear_brute_force(client_ip)
    return {"ok": True}


@app.get("/api/health")
async def api_health():
    return {
        "status": "ok",
        "service": "Saudiya Biletlar backend faol ishlayapti",
        "build": APP_BUILD,
    }


@app.get("/api/version")
async def api_version():
    """Qaysi build jonli ishlayotganini tekshirish uchun (eski deployni aniqlash)."""
    return {
        "build": APP_BUILD,
        "features": APP_BUILD_FEATURES,
        "server_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


@app.get("/api/payment-info")
async def api_payment_info():
    return {
        "card_number": settings.PAYMENT_CARD_NUMBER,
        "card_owner": settings.PAYMENT_CARD_OWNER,
        "admin_username": settings.ADMIN_USERNAME,
    }


# ==================== MINI APP FRONTEND STATIC MOUNT ====================
if os.path.exists("static"):
    app.mount("/", StaticFiles(directory="static", html=True), name="static")
